from flask import request
from flask import redirect
from flask import Flask, render_template
from flask import *
import pyrebase
import pandas as pd
import numpy as np
from xlrd import open_workbook
from flask_bootstrap import Bootstrap
import shutil
from pathlib import Path
from os import path
#from datetime import datetime
from flask_jsglue import JSGlue
from flask_mail import *  
from flask import Flask
from flask_mail import Mail, Message
import datetime
import openpyxl
import string
import random

wb=open_workbook('/Users/lee/Desktop/data.xlsx')

data=pd.read_excel(wb)
data['데이터 기준 일자']=pd.to_datetime(data['데이터 기준 일자'], format='%Y%m%d')

date_list=data['데이터 기준 일자'].unique()
dist=[]
total_dist=[]

def find_odometer(dt):
  find=dt['주행 기록계 값']
  find=round(find, 3)
  total_dist.append(find)

def find_daily_distance(dt):
  find=dt['기록계 값 차이 (전일 기준) (km)']
  find=round(find,3)
  dist.append(find)

def div_data(dt): # div_data(data_list)

  for i in range(len(dt)):
    date=dt[i]
    n_data=data[data['데이터 기준 일자']==date]
    n_dist=n_data.iloc[-1]
    find_odometer(n_dist) 
    find_daily_distance(n_dist)

div_data(date_list) # total_dist, dist 생성



def with_n_data(dl): # with_n_data(date_list)
  harsh=0

  for i in range(len(dl)):
    sum=0
    date_dist=dist[i]
    date=dl[i]   
    n_data=data[data['데이터 기준 일자']==date]
    sum=n_data['가속(150-200km/h)'].sum()

    if sum >= (0.2*date_dist): # 일일 주행거리 기준 0.2배 이상 가속주행 확인
      harsh += 1
  
  return harsh

def pressure(dt): # pressure(data)
  MAX_Pres=44
  NORM_Pres=32
  
  harsh=0
  count=0
  pres=0

  first_pres=dt['타이어 압력'].iloc[0]

  for i in range(1, dt.shape[0]):
    pres=dt['타이어 압력'].iloc[i]
    if pres is not first_pres:
      count += 1

  if (count/len(date_list)) >= 0.3:
    harsh += 1
  
  return harsh

def driven_distance(dt): # 20km 이하로 주 3-4회 이상 주행하면 가혹조건 +1 // driven_distance(dist)

  AVG_DIST=20
  NUM_DIST=3 
  harsh=0
  avg_dist=np.mean(dt)
  num_dist=len(dt)
  count=0

  for i in range(len(dt)):
    check=dt[i]
    if check <= AVG_DIST:
      count+=1

  if count>= NUM_DIST:
    harsh=harsh+1
    
  return harsh

def Harsh_condition(db):
  h1=with_n_data(date_list)
  h2=pressure(db)
  h3=driven_distance(dist)
  harsh=h1+h2+h3
  
  if harsh >= 2:
    state='가혹 조건'

  else:
    state='통상 조건'

  return state

state=Harsh_condition(data)

config = {
    "apiKey": "AIzaSyBiOhSK4jeNXyhA4vN4ebS9POw8D-1xunk",
    "authDomain": "newcar-6b31f.firebaseapp.com",
    "databaseURL": "https://newcar-6b31f.firebaseio.com/",
    "projectId": "newcar-6b31f",
    "storageBucket": "newcar-6b31f.appspot.com",
    "messagingSenderId": "220846431758",
    "appId": "1:220846431758:web:4e551b0d9ed469000ca65f",
    "measurementId": "G-Q5ER9J7EDM"
  }

firebase = pyrebase.initialize_app(config)
db = firebase.database()
jsglue = JSGlue()









app = Flask(__name__)

Bootstrap(app)
jsglue.init_app(app)

@app.route('/realmain', methods=['POST', 'GET'])
def realmain():
    if request.method == 'POST':
      global id
      global pwd
      global address
      global state

      id = request.form['id']
      pwd = request.form['pwd']
      address = request.form['address']
      necessaryAgree1 = request.form['necessaryAgree1']
      necessaryAgree2 = request.form['necessaryAgree2']
      selectAgree1 = request.form['selectAgree1']
      selectAgree2 = request.form['selectAgree2']

      if necessaryAgree1 == 'no' or necessaryAgree2 == 'no':
        return render_template('index.html')
      else:
        db.child('user').child(id).set({"id": id, "pwd": pwd, "address": address, "state": state, "necessaryAgree1": necessaryAgree1,"selectAgree1": selectAgree1, "necessaryAgree2": necessaryAgree2,"selectAgree2": selectAgree2 })
        return render_template('realmain.html', id = id, pwd = pwd, state = state)
    return render_template('index.html')

@app.route('/', methods=['POST', 'GET'])
def main():
  global id
  global pwd
  global address
  global state
  return render_template('realmain.html')

@app.route('/gogo', methods=['POST', 'GET'])
def gogo():
  if request.method == 'POST':
    now = datetime.datetime.today()
    userDate = db.child('user').child(id2).child('date').get().val()
    userFirstDate = db.child('user').child(id2).child('firstDate').get().val()
    userKilo = db.child('user').child(id2).child('kilometer').get().val()
    userGetKilo = db.child('user').child(id2).child('getKilo').get().val()
    return render_template('mainpage.html', admin_id = userId , state = state, now = now , firstDate = userFirstDate, date=userDate, kilo = userKilo, getKilo = userGetKilo)
  return render_template('mainpage.html')

@app.route('/about')
def about():
  return render_template('about.html')

@app.route('/contact')
def contact():
  return render_template('contact.html')

# @app.route('/logout')
# def logout():
#   if request.method == 'POST':
#     return render_template('realmain.html')
#   return render_template('realmain.html')

@app.route('/about2',methods=['POST', 'GET'])
def about2():
  if request.method == 'POST':
    now = datetime.datetime.today()
    userDate = db.child('user').child(id2).child('date').get().val()
    userFirstDate = db.child('user').child(id2).child('firstDate').get().val()
    userKilo = db.child('user').child(id2).child('kilometer').get().val()
    userGetKilo = db.child('user').child(id2).child('getKilo').get().val()
    return render_template('about2.html', admin_id = userId , state = state, now = now , firstDate = userFirstDate, date=userDate, kilo = userKilo, getKilo = userGetKilo)
  return render_template('about2.html')

@app.route('/contact2',methods=['POST', 'GET'])
def contact2():
  if request.method == 'POST':
    now = datetime.datetime.today()
    userDate = db.child('user').child(id2).child('date').get().val()
    userFirstDate = db.child('user').child(id2).child('firstDate').get().val()
    userKilo = db.child('user').child(id2).child('kilometer').get().val()
    userGetKilo = db.child('user').child(id2).child('getKilo').get().val()
    return render_template('contact2.html', admin_id = userId , state = state, now = now , firstDate = userFirstDate, date=userDate, kilo = userKilo, getKilo = userGetKilo)
  return render_template('contact2.html')


@app.route('/mydata', methods=['POST', 'GET'])
def mydata():
  if request.method == 'POST':
    selectAgree1 = request.form['selectAgree1']
    selectAgree2 = request.form['selectAgree2']
    necessaryAgree1 = request.form['necessaryAgree1']
    necessaryAgree2 = request.form['necessaryAgree2']
    db.child('user').child(id2).update({"selectAgree1":selectAgree1,"selectAgree2":selectAgree2, "necessaryAgree1":necessaryAgree1,"necessaryAgree2":necessaryAgree2})
    now = datetime.datetime.today()
    userDate = db.child('user').child(id2).child('date').get().val()
    userFirstDate = db.child('user').child(id2).child('firstDate').get().val()
    userKilo = db.child('user').child(id2).child('kilometer').get().val()
    userGetKilo = db.child('user').child(id2).child('getKilo').get().val()
    if necessaryAgree1 == 'no' or necessaryAgree2 == 'no':
      db.child('user').child(id2).set({"삭제된 아이디입니다": "."})
      return render_template('login.html')
  return render_template('mainpage.html', admin_id = userId , state = state, now = now , firstDate = userFirstDate, date=userDate, kilo = userKilo, getKilo = userGetKilo)

@app.route('/purchase', methods=['POST', 'GET'])
def purchase():
  now = datetime.datetime.today()
  userDate = db.child('user').child(id2).child('date').get().val()
  userFirstDate = db.child('user').child(id2).child('firstDate').get().val()
  userKilo = db.child('user').child(id2).child('kilometer').get().val()
  userGetKilo = db.child('user').child(id2).child('getKilo').get().val()
  return render_template('purchase.html',  admin_id = userId , state = state, now = now , firstDate = userFirstDate, date=userDate, kilo = userKilo, getKilo = userGetKilo)

@app.route('/login', methods=['POST', 'GET'])
def login():
  if request.method == 'POST':
    global id2
    global pwd2
    global userId
    global userPwd
    id2 = request.form['id2']
    pwd2 = request.form['pwd2']
    userId = db.child('user').child(id2).child('id').get().val()
    userPwd = db.child('user').child(id2).child('pwd').get().val()
    if id2 == userId and pwd2 == userPwd:
      global date
      global kilo
      global getKilo
      global firstDate
      global now
      now = datetime.datetime.today()
      userDate = db.child('user').child(id2).child('date').get().val()
      userFirstDate = db.child('user').child(id2).child('firstDate').get().val()
      userKilo = db.child('user').child(id2).child('kilometer').get().val()
      userGetKilo = db.child('user').child(id2).child('getKilo').get().val()

      return render_template('mainpage.html', admin_id = userId , state = state, now = now , firstDate = userFirstDate, date=userDate, kilo = userKilo, getKilo = userGetKilo)
    else:
      return render_template('login.html')
  return render_template('login.html')

@app.route('/mainpage')
def mainpage():
  return render_template('mainpage.html',admin_id = userId, state = state )

@app.route('/upload', methods=['POST', 'GET'])
def upload():
  global date
  global kilo
  global getKilo
  global firstDate
  global now
  userDate = db.child('user').child(id2).child('date').get().val()
  userFirstDate = db.child('user').child(id2).child('firstDate').get().val()
  userKilo = db.child('user').child(id2).child('kilometer').get().val()
  userGetKilo = db.child('user').child(id2).child('getKilo').get().val()

  if request.method == 'POST':
    date = request.form['date']
    firstDate = request.form['firstDate']
    kilo = request.form['kilo']
    getKilo = request.form['getKilo']
    cartype = request.form['cartype']
    caroil = request.form['caroil']
    db.child('user').child(id2).update({"date": date, "firstDate": firstDate, "kilometer": kilo, 'getKilo' : getKilo, 'cartype': cartype, 'caroil': caroil})

    userDate = db.child('user').child(id2).child('date').get().val()
    userFirstDate = db.child('user').child(id2).child('firstDate').get().val()
    userKilo = db.child('user').child(id2).child('kilometer').get().val()
    userGetKilo = db.child('user').child(id2).child('getKilo').get().val()
    now = datetime.datetime.today()
    # firstDate = firstDate,  now=now, date=date, kilo = kilo
    return render_template('mainpage.html', admin_id = userId, state = state, firstDate = userFirstDate,  now=now, date=userDate, kilo = userKilo, getKilo = userGetKilo)
  return render_template('mainpage.html', admin_id = userId, state = state, firstDate = userFirstDate,  now=now, date=userDate, kilo = userKilo, getKilo = userGetKilo)

@app.route('/pay')
def pay():
  return render_template('pay.html',admin_id = userId, state = state )



@app.route("/getPlotCSV")
def getPlotCSV():
    # with open("outputs/Adjacency.csv") as fp:
    #     csv = fp.read()
    csv = '1,2,3\n4,5,6\n'
    return Response(
        csv,
        mimetype="text/csv",
        headers={"Content-disposition":
                 "attachment; filename=myplot.csv"})

@app.route('/plot_csv') # this is a job for GET, not POST
def plot_csv():
    now = datetime.datetime.now()
    nowDatetime = now.strftime('%Y-%m-%d %H:%M:%S')
    length = 8
    string_pool = string.digits
    DownloadId = ""
    for i in range(length):
      DownloadId += random.choice(string_pool)
    DownloadWay = '엑셀파일다운로드'
    DownloadContents = '차량 주행 데이터'
    DownloadForm = 'xlsx'
    DownloadIssuer = '카링(CARing)'
    DownloadManager = '김도현'
    db.child('user').child(id2).child('데이터영수증_다운로드').update({
      "DownloadDate": nowDatetime, 
      "DownloadId": DownloadId,
      "DownloadWay": DownloadWay,
      "DownloadContents": DownloadContents,
      "DownloadForm": DownloadForm,
      "DownloadIssuer": DownloadIssuer,
      "DownloadManager": DownloadManager })
    return render_template('index.html')
    # return send_file('/Users/lee/Desktop/data.xlsx',
    #                  mimetype='text/xlsx',
    #                  attachment_filename='Mydata.xlsx',
    #                  as_attachment=True)

# html로 데이터 받기 메일보내기!@#!@#!@#!@#@!#@!#@!
@app.route('/plot_email')  
def plot_email():  
    now = datetime.datetime.now()
    nowDatetime = now.strftime('%Y.%m.%d..%H.%M.%S')
    length = 8
    string_pool = string.digits
    DownloadId = ""
    for i in range(length):
      DownloadId += random.choice(string_pool)
    DownloadWay = 'html파일'
    DownloadContents = '차량 주행 데이터'
    DownloadForm = 'html'
    DownloadIssuer = '카링(CARing)'
    DownloadManager = '김도현'
    db.child('user').child(id2).child('데이터영수증_이메일').update({
      "DownloadDate": nowDatetime, 
      "DownloadId": DownloadId,
      "DownloadWay": DownloadWay,
      "DownloadContents": DownloadContents,
      "DownloadForm": DownloadForm,
      "DownloadIssuer": DownloadIssuer,
      "DownloadManager": DownloadManager })


    wbb = pd.read_excel('/Users/lee/Desktop/data.xlsx')
    wbb.to_html('/Users/lee/Desktop/firstProject/public/templates/data.html')
    msg = Message('subject', sender = 'hohooodo@gmail.com', recipients=['helloweew2345@gmail.com'])  
    msg.html = render_template('data.html')

    with app.open_resource("/Users/lee/Desktop/data.xlsx") as fp:
      msg.attach("data.xlsx", "application/vnd.ms-excel", fp.read())
    mail.send(msg)
    return "Mail Sent, Please check the mail id"  




@app.route('/dataReceipt',methods=['POST', 'GET'])
def dataReceipt():
  if request.method == 'POST':
    typeOfr = request.form['typeOfr']
    if typeOfr == 'download':
      return send_file('/Users/lee/Desktop/firstProject/public/data_receipt.xlsx',
                        mimetype='text/xml',
                        attachment_filename='Data_Receipt.xlsx',
                        as_attachment=True)
    elif typeOfr == 'agree':
      # now = datetime.datetime.now()
      now = datetime.datetime.today()
      nowDatetime = now.strftime('%Y-%m-%d %H:%M:%S')
      length = 10
      string_pool = string.digits
      DownloadId = ""
      for i in range(length):
        DownloadId += random.choice(string_pool)
      DownloadIssuer = '카링(CARing)'
      DownloadManager = '김도현'
      wwbb = openpyxl.Workbook()
      sheet = wwbb.active
    # wwbb.create_sheet(title=nowDatetime)
      sheet['A1'] = '데이터 영수증'
      # select = db.child('user').child(id2).child('selectAgree').get().val()
      sheet.cell(row=2, column=1).value = '영수증 종류'
      sheet.cell(row=2, column=4).value = '동의내역'
      sheet.cell(row=3, column=1).value = '필수정보동의'
      sheet.cell(row=3, column=1).value = '영수증 ID'
      sheet.cell(row=3, column=4).value = DownloadId
      sheet.cell(row=4, column=1).value = '내려받기 시간'
      sheet.cell(row=4, column=4).value = nowDatetime
      sheet.cell(row=5, column=1).value = '필수정보동의'
      sheet.cell(row=5, column=4).value = db.child('user').child(id2).child('necessaryAgree').get().val()
      sheet.cell(row=6, column=1).value = '선택정보동의'
      sheet.cell(row=6, column=4).value = db.child('user').child(id2).child('selectAgree').get().val()
      sheet.cell(row=7, column=1).value = '발행기관'
      sheet.cell(row=7, column=4).value = DownloadIssuer
      sheet.cell(row=8, column=1).value = '담당자'
      sheet.cell(row=8, column=4).value = DownloadManager
      wwbb.save('data_receipt_agree.xlsx')
      return send_file('/Users/lee/Desktop/firstProject/public/data_receipt_agree.xlsx',
                        mimetype='text/xml',
                        attachment_filename='Data_Receipt_agree.xlsx',
                        as_attachment=True)


  return render_template('dataReceipt.html')

  # select = db.child('user').child(id2).child('selectAgree').get().val()
  # necessary = db.child('user').child(id2).child('necessaryAgree').get().val()

#Flask mail configuration  
app.config['MAIL_SERVER']='smtp.gmail.com'  
app.config['MAIL_PORT']=123123123
app.config['MAIL_USERNAME'] = 'example@gmail.com'  
app.config['MAIL_PASSWORD'] = 'example'  
app.config['MAIL_USE_TLS'] = False  
app.config['MAIL_USE_SSL'] = True  
#instantiate the Mail class  
mail = Mail(app)  
@app.route('/dataDownload' ,methods=['POST', 'GET'])
def dataDownload():
  if request.method == 'POST':
    userDate = db.child('user').child(id2).child('date').get().val()
    userFirstDate = db.child('user').child(id2).child('firstDate').get().val()
    userKilo = db.child('user').child(id2).child('kilometer').get().val()
    userGetKilo = db.child('user').child(id2).child('getKilo').get().val()
    kindOfData = request.form['n1']
    dataForm = request.form['n2']
    downloadForm = request.form['n3']
    now = datetime.datetime.now()
    #now = datetime.today()
    nowDatetime = now.strftime('%Y-%m-%d %H:%M:%S')
    length = 10
    string_pool = string.digits
    DownloadId = ""
    for i in range(length):
      DownloadId += random.choice(string_pool)
    DownloadWay = dataForm
    DownloadContents = kindOfData
    DownloadIssuer = '카링(CARing)'
    DownloadManager = '김도현'
    db.child('user').child(id2).child('데이터영수증').update({
      "DownloadDate": nowDatetime, 
      "DownloadId": DownloadId,
      "DownloadWay": DownloadWay,
      "DownloadContents": DownloadContents,
      "DownloadIssuer": DownloadIssuer,
      "DownloadManager": DownloadManager })

    wwbb = openpyxl.Workbook()
    sheet = wwbb.active
    # wwbb.create_sheet(title=nowDatetime)
    sheet['A1'] = '데이터 영수증'
    sheet.cell(row=2, column=1).value = '영수증 종류'
    sheet.cell(row=2, column=4).value = '내려받기 영수증'
    sheet.cell(row=3, column=1).value = '영수증 ID'
    sheet.cell(row=3, column=4).value = DownloadId
    sheet.cell(row=4, column=1).value = '내려받기 시간'
    sheet.cell(row=4, column=4).value = nowDatetime
    sheet.cell(row=5, column=1).value = '내려받기 형식'
    sheet.cell(row=5, column=4).value = DownloadWay
    sheet.cell(row=6, column=1).value = '다운로드 내역'
    sheet.cell(row=6, column=4).value = DownloadContents
    sheet.cell(row=7, column=1).value = '발행기관'
    sheet.cell(row=7, column=4).value = DownloadIssuer
    sheet.cell(row=8, column=1).value = '담당자'
    sheet.cell(row=8, column=4).value = DownloadManager
    wwbb.save('data_receipt.xlsx')
    if kindOfData == '차량운행 데이터' and dataForm == 'csv' and downloadForm == 'deviceD':
      return send_file('/Users/lee/Desktop/data.csv',
                        mimetype='text/plain',
                        attachment_filename='Mydata.csv',
                        as_attachment=True)

    elif kindOfData == '차량운행 데이터' and dataForm =='xlsx' and downloadForm == 'deviceD':
      return send_file('/Users/lee/Desktop/data.xlsx',
                        mimetype='text/xml',
                        attachment_filename='Mydata.xlsx',
                        as_attachment=True)

    elif kindOfData == '차량운행 데이터' and dataForm =='HTML' and downloadForm == 'deviceD':
      return send_file('/Users/lee/Desktop/data.html',
                        mimetype='text/html',
                        attachment_filename='Mydata.html',
                        as_attachment=True)
    if kindOfData == '사용자 운행패턴 데이터' and dataForm == 'csv' and downloadForm == 'deviceD':
      return send_file('/Users/lee/Desktop/data.csv',
                        mimetype='text/plain',
                        attachment_filename='Mydata.csv',
                        as_attachment=True)

    elif kindOfData == '사용자 운행패턴 데이터' and dataForm =='xlsx' and downloadForm == 'deviceD':
      return send_file('/Users/lee/Desktop/data2.xlsx',
                        mimetype='text/xml',
                        attachment_filename='Mydata.xlsx',
                        as_attachment=True)

    elif kindOfData == '사용자 운행패턴 데이터' and dataForm =='HTML' and downloadForm == 'deviceD':
      return send_file('/Users/lee/Desktop/data2.html',
                        mimetype='text/html',
                        attachment_filename='Mydata.html',
                        as_attachment=True)

# 이메일 다운로드
    if kindOfData == '차량운행 데이터' and dataForm == 'csv' and downloadForm == 'emailD':
      msg = Message('subject', sender = 'hohooodo@gmail.com', recipients=['helloweew2345@gmail.com'])  
      with app.open_resource("/Users/lee/Desktop/data.csv") as fp:
        msg.attach("data.csv", "txt/plain", fp.read())
        mail.send(msg)

    elif kindOfData == '차량운행 데이터' and dataForm =='xlsx' and downloadForm == 'emailD':
      msg = Message('subject', sender = 'hohooodo@gmail.com', recipients=['helloweew2345@gmail.com'])  
      with app.open_resource("/Users/lee/Desktop/data.xlsx") as fp:
        msg.attach("data.xlsx", "application/vnd.ms-excel", fp.read())
        mail.send(msg)

    elif kindOfData == '차량운행 데이터' and dataForm =='HTML' and downloadForm == 'emailD':
      msg = Message('subject', sender = 'hohooodo@gmail.com', recipients=['helloweew2345@gmail.com'])  
      with app.open_resource("/Users/lee/Desktop/data.html") as fp:
        msg.attach("data.html", "text/html", fp.read())
        mail.send(msg)

    if kindOfData == '사용자 운행패턴 데이터' and dataForm == 'csv' and downloadForm == 'emailD':
      msg = Message('subject', sender = 'hohooodo@gmail.com', recipients=['helloweew2345@gmail.com'])  
      with app.open_resource("/Users/lee/Desktop/data2.csv") as fp:
        msg.attach("data2.csv", "txt/plain", fp.read())
        mail.send(msg)

    elif kindOfData == '사용자 운행패턴 데이터' and dataForm =='xlsx' and downloadForm == 'emailD':
      msg = Message('subject', sender = 'hohooodo@gmail.com', recipients=['helloweew2345@gmail.com'])  
      with app.open_resource("/Users/lee/Desktop/data2.xlsx") as fp:
        msg.attach("data2.xlsx", "application/vnd.ms-excel", fp.read())
        mail.send(msg)

    elif kindOfData == '사용자 운행패턴 데이터' and dataForm =='HTML' and downloadForm == 'emailD':
      msg = Message('subject', sender = 'hohooodo@gmail.com', recipients=['helloweew2345@gmail.com'])  
      with app.open_resource("/Users/lee/Desktop/data2.html") as fp:
        msg.attach("data2.html", "text/html", fp.read())
        mail.send(msg)


  # global select
  # global necessary
  # select = db.child('user').child(id2).child('selectAgree').get().val()
  # necessary = db.child('user').child(id2).child('necessaryAgree').get().val()

  return render_template('mainpage.html',  admin_id = userId, state = state, firstDate = userFirstDate,  now=now, date=userDate, kilo = userKilo, getKilo = userGetKilo)






@app.route('/aa')
def aa():

  return render_template('data.html')






















if __name__ == '__main__':
    app.run(debug=True)



#  db.child('user').child(id2).update({"selectAgree":selectAgree, "necessaryAgree":necessaryAgree})
#     #vv = db.child('user').child(id2).child('v1').get().val()
#     if necessaryAgree == 'no':
#       db.child('user').child(id2).set({"삭제된 아이디입니다": "."})
#       return render_template('login.html')


